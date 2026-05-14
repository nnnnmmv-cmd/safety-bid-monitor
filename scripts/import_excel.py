"""엑셀 명부(건설공사안전점검 명부.xlsx)를 Supabase sites 테이블로 일괄 등록.

사용 예:
    python scripts/import_excel.py --file ~/Downloads/건설공사안전점검\\ 명부.xlsx
    python scripts/import_excel.py --file ... --delete-examples   # 예시-* 사전 삭제
    python scripts/import_excel.py --file ... --dry-run            # 적용 안 하고 미리보기
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

import openpyxl
from dotenv import load_dotenv

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src import store
from src.config import PROJECT_ROOT

load_dotenv(PROJECT_ROOT / ".env")


# 엑셀 컬럼 → Supabase sites 필드 매핑 (1-based 인덱스)
COL_MAP: dict[int, str] = {
    3:  "name",
    4:  "category",
    5:  "homecheck",
    6:  "hansijin",
    7:  "hanjugum",
    8:  "bidding_status",
    9:  "new_submission_date",
    10: "period_start",
    11: "period_end",
    13: "announce_planned_date",
    14: "previous_announce_date",
    15: "previous_deadline",
    16: "under_100m_winner_method",
    17: "above_100m_winner_method",
    18: "bid_submission_method",
    19: "performance_proof",
    20: "work_overlap_doc",
}

DATE_FIELDS: set[str] = {
    "new_submission_date", "period_start", "period_end",
    "announce_planned_date", "previous_announce_date", "previous_deadline",
}

CATEGORY_NORMALIZE: dict[str, str] = {
    "토목/건축": "건축·토목",
    "건축/토목": "건축·토목",
    "토목·건축": "건축·토목",
}


def _norm_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return _norm_str(value)[:10]


def _norm_category(value: Any) -> str:
    s = _norm_str(value)
    return CATEGORY_NORMALIZE.get(s, s)


def _row_to_site(row_values: list[Any], hyperlink_url: str = "") -> dict[str, Any] | None:
    name = _norm_str(row_values[2] if len(row_values) >= 3 else None)
    if not name:
        return None

    record: dict[str, Any] = {
        "name": name,
        "enabled": False,
        "region": "",
        "adapter": "egov",
        "base_url": "",
        "list_url": "",
        "list_params": {},
        "pagination": {"param": "pageIndex", "max_pages": 3},
        "selectors": {},
        "note": "",
    }
    for col_idx, field in COL_MAP.items():
        idx = col_idx - 1
        if idx >= len(row_values):
            continue
        raw = row_values[idx]
        if field == "category":
            record[field] = _norm_category(raw)
        elif field in DATE_FIELDS:
            record[field] = _norm_date(raw)
        else:
            record[field] = _norm_str(raw)

    # 지자체명 셀에 하이퍼링크가 있으면 URL을 base/list/params로 분해
    if hyperlink_url:
        p = urlparse(hyperlink_url)
        if p.scheme and p.netloc:
            record["base_url"] = f"{p.scheme}://{p.netloc}"
            record["list_url"] = f"{p.scheme}://{p.netloc}{p.path}"
            qs = parse_qs(p.query, keep_blank_values=False)
            record["list_params"] = {k: v[0] for k, v in qs.items() if v}
    return record


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="엑셀 파일 경로 (.xlsx)")
    parser.add_argument("--delete-examples", action="store_true", help="예시-* 발주청을 사전 삭제")
    parser.add_argument("--dry-run", action="store_true", help="실제 반영 없이 미리보기")
    args = parser.parse_args()

    xlsx_path = Path(args.file).expanduser()
    if not xlsx_path.exists():
        print(f"ERROR: 파일을 찾을 수 없습니다: {xlsx_path}")
        return 2

    if not store.is_configured():
        print("ERROR: Supabase 환경변수 미설정 — .env 확인")
        return 2

    # 하이퍼링크 추출하려면 read_only=False
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records: list[dict[str, Any]] = []
    with_url = 0
    skipped = 0
    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        name_cell = ws.cell(row=row_idx, column=3)
        url = ""
        if name_cell.hyperlink and name_cell.hyperlink.target:
            url = str(name_cell.hyperlink.target).strip()
        rec = _row_to_site(row_values, hyperlink_url=url)
        if rec is None:
            skipped += 1
            continue
        if rec.get("list_url"):
            with_url += 1
        records.append(rec)

    print(f"읽어들인 발주청: {len(records)}건 (빈 행 스킵: {skipped})")
    print(f"  URL 자동 등록: {with_url}건 / 빈 URL: {len(records) - with_url}건")

    # 동일 이름 중복은 카테고리 suffix를 붙여 분리 (예: '성남시' → '성남시-건축', '성남시-토목')
    groups: dict[str, list[dict[str, Any]]] = {}
    for r in records:
        groups.setdefault(r["name"], []).append(r)

    final_records: list[dict[str, Any]] = []
    conflict_originals: list[str] = []
    renamed: list[tuple[str, str]] = []
    for name, rows in groups.items():
        if len(rows) == 1:
            final_records.append(rows[0])
            continue
        conflict_originals.append(name)
        for i, r in enumerate(rows):
            cat = (r.get("category") or "").strip()
            suffix = f"-{cat}" if cat else f"-{i+1}"
            new_name = f"{name}{suffix}"
            r["name"] = new_name
            final_records.append(r)
            renamed.append((name, new_name))

    # suffix 적용 후에도 동명 발생 시 인덱스 추가
    by_name: dict[str, dict[str, Any]] = {}
    for r in final_records:
        if r["name"] in by_name:
            j = 2
            while f"{r['name']}#{j}" in by_name:
                j += 1
            r["name"] = f"{r['name']}#{j}"
        by_name[r["name"]] = r
    records = list(by_name.values())

    if conflict_originals:
        print(f"  ⚠ 중복 이름 {len(conflict_originals)}개 → 카테고리 suffix로 분리:")
        for original, new in renamed:
            print(f"      {original}  →  {new}")
    print(f"  최종 유니크 발주청: {len(records)}건")

    print("샘플 5건:")
    for r in records[:5]:
        url_short = r.get("list_url") or "(URL 없음)"
        if len(url_short) > 60:
            url_short = url_short[:57] + "..."
        print(f"  - {r['name']} | {r.get('category','')} | {url_short}")

    if args.dry_run:
        print("\n--dry-run 모드 — Supabase에 적용하지 않음.")
        return 0

    existing = store.list_sites()
    existing_names = {s.get("name") for s in existing}

    if args.delete_examples:
        to_delete = [n for n in existing_names if n and n.startswith("예시-")]
        for n in to_delete:
            store.delete_site(n)
        if to_delete:
            print(f"\n예시-* {len(to_delete)}건 삭제: {to_delete}")

    # 카테고리 분리된 발주청의 기존 단일 항목 (예: 기존에 '성남시' 단독 등록)을 정리
    if conflict_originals:
        cleaned_originals = [n for n in conflict_originals if n in existing_names]
        for n in cleaned_originals:
            store.delete_site(n)
        if cleaned_originals:
            print(f"기존 단일 항목 {len(cleaned_originals)}건 정리 (suffix 버전으로 대체): {cleaned_originals}")

    print(f"\nSupabase에 upsert 중... ({len(records)}건)")
    store.upsert_sites(records)
    print("=== 완료 ===")
    after = store.list_sites()
    print(f"현재 sites 총 {len(after)}건")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
